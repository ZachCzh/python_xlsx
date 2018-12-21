# xls_to_xlsx

# change the working directory
import os
while True:
    cwd = os.getcwd() 
    wd_ch = input("Working on "+cwd+", do you wish to change? Y/N:")

    if wd_ch == 'Y' or wd_ch == 'y' or wd_ch == 'Yes' or wd_ch == 'yes' :
        while True:
            cwd_n = input("Please input the folder directory:")
            try:
                os.chdir(cwd_n) 
                break
            except FileNotFoundError or OSError:
                print("Error! Please input a valid directory, e.g.'f:/myfolder'")
                continue
        print(f"Successfully changed to {cwd_n}")
        break
    elif wd_ch == 'N' or wd_ch == 'n' or wd_ch == 'No' or wd_ch == 'no' :
        cwd_n = cwd
        break
    else:
        print("Error! Please answer 'Yes' or 'No'")
        continue

# show all the files in working directory
files = []
print('\n'+"This folder contains following files:"+'\n')
for filename in os.listdir(cwd_n):
    print(filename)
    files.append(filename)

# create a new excel file
import openpyxl
print('\n'+"Creating a new excel file"+'\n')

n_wb = openpyxl.Workbook()
n_ws = n_wb.active

# open the .xls file
import xlrd
ma_filename = input('\n'+"Please input the excel file name (input 'all' to read all):")
if ma_filename == 'all':
    row = 1
    for filename in files:
        wb = xlrd.open_workbook(filename)
        sheet = wb.sheets()[0]
        name = sheet.cell(3,3).value
        ma_name = name[5:]
        use_rate = sheet.cell(37,3).value

        n_ws.cell(row = row, column = 1, value = ma_name)
        n_ws.cell(row = row, column = 2, value = use_rate)

        row = row + 1

# acquire today's date
import datetime
now_date = datetime.datetime.today().strftime('%Y%m%d')

# save as a new excel file
while True:
    wb_name = input("Do you wish to save this workbook? Y/N:")
    if wb_name == 'Y' or wb_name == 'y' or wb_name == 'Yes' or wb_name == 'yes':
        while True:
            try:
                file_name = input("Please input the new file name:")
                break
            except ValueError:
                print("Error! Please input at least one character")
                continue
        new_name = file_name + now_date + ".xlsx"
        n_wb.save(new_name)
        print(f"Successfully saved excel file {file_name + now_date}")
        break
    elif wb_name == 'N' or wb_name == 'n' or wb_name == 'No' or wb_name == 'no':
        break  
    else:
        print("Error! Please answer 'Yes' or 'No'")
        continue
