# xlsx_ne.py

from __main__ import files
import openpyxl

# read the excel file needed to work on
while True:
    xl_name = input('\n'+"Please input the excel file name:")
    try:
        wb = openpyxl.load_workbook(xl_name)
        break
    except FileNotFoundError:
        print("Error! Please input a valid file name with suffix '.xlsx'")
        continue
print (f"This workbook contains {wb.sheetnames}")

# read the worksheet needed to work on
while True:
    sn = input("Please input the number of sheetname needed to be editted:")
    try:
        ws = wb[wb.sheetnames[int(sn)-1]]
        break
    except ValueError or IndexError:
        print("Error! Please input a valid number of sheetname")
        continue

# count the total number of row
nrows = ws.max_row
# count the total number of column
ncols = ws.max_column

# create a new excel file
print('\n'+"Creating a new excel file"+'\n')

n_wb = openpyxl.Workbook() 

# rename the active sheet
while True:
    sheet = n_wb.active
    sn_ch = input("Active worksheet named '"+sheet.title+"', do you wish to rename it? Y/N:")
    if sn_ch == 'Y' or sn_ch == 'y' or sn_ch == 'Yes' or sn_ch == 'yes':
        sn_new = input("Please input the new name:")
        sheet.title = sn_new
        print(f"Successfully change sheetname to {sheet.title}")
        break
    elif sn_ch == 'N' or sn_ch == 'n' or sn_ch == 'No' or sn_ch == 'no':
        sn_new = sheet
        break
    else:
        print("Error! Please answer 'Yes' or 'No'")
        continue
print (f"This workbook contains {n_wb.sheetnames}")

# copy all data into the new excel
for row in range(1, nrows+1):
    for col in range(1, ncols+1):
        sn_new.cell(row = row, column = col, value = ws.cell(row=row,column=col).value)

# acquire today's date
import datetime
now_date = datetime.datetime.today().strftime('%Y%m%d')

# save as a new excel file
while True:
    wb_name = input("Do you wish to save this workbook? Y/N:")
    if wb_name == 'Y' or wb_name == 'y' or wb_name == 'Yes' or wb_name == 'yes':
        while True:
            try:
                file_name = input("Please input the new file name:")
                break
            except ValueError:
                print("Error! Please input at least one character")
                continue
        new_name = file_name + now_date + ".xlsx"
        n_wb.save(new_name)
        print(f"Successfully saved excel file {file_name + now_date}")
        break
    elif wb_name == 'N' or wb_name == 'n' or wb_name == 'No' or wb_name == 'no':
        break  
    else:
        print("Error! Please answer 'Yes' or 'No'")
        continue
